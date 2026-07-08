"""
Ten moduł generuje plik Excel (.xlsx) z listą kandydatów.

Ważna decyzja projektowa: baza SQLite jest "źródłem prawdy" o wszystkich
kandydatach. Za każdym razem, gdy pojawia się nowy kandydat, cały plik
Excel jest generowany OD NOWA na podstawie danych z bazy - dzięki temu:
  - nigdy nie tracimy poprzednich wpisów (są bezpiecznie w bazie),
  - lista jest zawsze poprawnie posortowana malejąco wg oceny,
  - nie ma ryzyka uszkodzenia pliku przez częściowy zapis.
"""

import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger("cv_bot")

COLUMNS = [
    ("Imię i nazwisko", "imie_nazwisko", 25),
    ("E-mail", "email", 28),
    ("Telefon", "telefon", 16),
    ("Ocena (0-100)", "ocena", 14),
    ("Status", "status", 14),
    ("Uzasadnienie oceny", "uzasadnienie", 60),
    ("Stanowisko", "stanowisko", 25),
    ("Data otrzymania", "data_otrzymania", 22),
    ("Ścieżka do pliku CV", "sciezka_pliku", 45),
]


def write_excel(candidates: list, output_path: str) -> None:
    """Zapisuje listę kandydatów (posortowaną malejąco wg oceny) do pliku Excel.
    `candidates` to lista słowników, np. wynik z database.get_all_candidates()."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kandydaci"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # Kandydaci już przychodzą posortowani z bazy (ORDER BY ocena DESC),
    # ale sortujemy jeszcze raz na wszelki wypadek.
    sorted_candidates = sorted(candidates, key=lambda c: c.get("ocena") or 0, reverse=True)

    for row_idx, candidate in enumerate(sorted_candidates, start=2):
        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            value = candidate.get(field, "")
            if field == "status":
                value = "Rezerwowy" if value == "rezerwowy" else "Aktywny"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "uzasadnienie"))

            # Koloruj oceny: zielone wysokie, żółte średnie, czerwone niskie
            if field == "ocena" and isinstance(value, int):
                if value >= 70:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value >= 40:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb.save(output_path)
    logger.info("Zapisano plik Excel: %s (%d kandydatów)", output_path, len(sorted_candidates))
